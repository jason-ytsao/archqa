#!python
# jasonx.tsao@intel.com - [09/30/2022]

import re
import numpy as np
from pathlib import Path
import pandas as pd
import argparse as aps
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import logging

"""
1. Customized for HW architecture features comparison between 
    - `'CSME IE OCS Hardware Architecture Features Per Project.xlsm'` and 
    -  HW arch features config sheet from ArchGUI database
2. Comparison between two HW architecture features config sheets from ArchGUI database
3. To compare two sheets (*.xlsx)
4. To format HW architecture features config sheet from ArchGUI database
"""
class XlsxDiff:

    def __init__(self, out_dir: str, mode: str):
        self.df_extract_projs_start = 0
        self.mode = mode
        self.database_projIPs = []
        self.database_projs = []
        self.database_ips = []
        self.database_df_extract_BF = []
        pd.set_option('display.max_rows', 500)
        
        # OUTPUT FILES:
        self.cwd = Path.cwd()
        self.output_dir = out_dir
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        self.wr_discrepancy = f"{Path(self.output_dir, 'output_diffcfg.xlsx')}"
        self.wr_matched = f"{Path(self.output_dir, 'output_matched.xlsx')}"
        self.wr_diffBF_all = f"{Path(self.output_dir, 'output_diffbf.xlsx')}"
        self.wr_diffSS = f"{Path(self.output_dir, 'output_diffss.xlsx')}"
        if mode == 'diffgd':
            self.wr_database_extract_sorted = f"{Path(self.output_dir, 'output_extract_sorted_db.xlsx')}"
            self.wr_golden_extract_sorted = f"{Path(self.output_dir, 'output_extract_sorted_golden.xlsx')}"
            self.wr_diff_golden = f"{Path(self.output_dir, 'output_diffcfg_golden.xlsx')}"
            self.wr_diff_database = f"{Path(self.output_dir, 'output_diffcfg_db.xlsx')}"
            self.wr_diffBF_golden = f"{Path(self.output_dir, 'output_diffbf_extragolden.xlsx')}"
            self.wr_diffBF_database = f"{Path(self.output_dir, 'output_diffbf_extradb.xlsx')}"
            self.wr_waived_golden = f"{Path(self.output_dir, 'output_waived_golden.xlsx')}"
        elif mode == 'diffdd':
            self.wr_database_extract_sorted = ''
            self.wr_diff_golden = ''
            self.wr_diff_database = ''
            self.wr_diffBF_golden = ''
            self.wr_diffBF_database = '' 


    def process_database(self, db: str) -> pd.DataFrame:
        """
        Process xlsx downloaded from ArchGUI
        """
        database_df = pd.read_excel(db)
        database_slice_start = list(database_df.columns).index('Type Name') + 1
        orig_proj_names = database_df.columns[database_slice_start:]

        logging.info(f"\nColumn Names of Proj and IP Generation = {database_df.columns[database_slice_start:].tolist()}")

        # [09/30/2022]
        # added for databse platform name changes 
        # [10/17/2022]
        # databse changed names again, modify names again accordingly 
        db_2_golden_name_mapping = {
            # 'PTL_PCD_P_CSME 4.11': 'PTL-PCD-P_CSME 4.11',       
            # 'PTL_PCD_P_ESE 2.1': 'PTL-PCD-P_ESE 2.1',
            # 'PTL_PCD_P_OSSE 1.1': 'PTL-PCD-P_OSSE 1.1',
            'PTL-CPU-P_GSC 3.3': 'PTL CD die_GSC 3.3'           # database : Golden sheet
        }
        # databse change proj naming from hyphens to underscores
        # Change db proj's names, replacing underscore with hyphen to match golden's naming
        def hyphen(name: str) -> str:
            if len(re.findall('_', name)) > 1:
                proj = '-'.join(name.split(' ')[0].split('_')[:-1])
                ip = name.split(' ')[0].split('_')[-1]
                version = name.split(' ')[-1]
                return f'{proj}_{ip} {version}'
            else: return name
        hyphen_proj_names = []
        for i in database_df.columns[database_slice_start:].tolist():
            hyphen_proj_names.append(hyphen(i))
        database_df.columns = database_df.columns[:database_slice_start].tolist() + hyphen_proj_names
        
        logging.info(f"Changing col names for mapping to Golden's...")
        for k, v in db_2_golden_name_mapping.items():
            if k in database_df.columns.tolist():
                logging.info(f"{k:<24} ===> {v:>24}")
        for i, j in zip(orig_proj_names, database_df.columns[database_slice_start:].tolist()):
            if i not in db_2_golden_name_mapping and i != j:
                logging.info(f"{i:<24} ==> {j:>24}")
        
        # Rename column labels to map Golden's
        # no errors even when column names is not found
        database_df.rename(
            columns={
                'PTL-CPU-P_GSC 3.3': db_2_golden_name_mapping['PTL-CPU-P_GSC 3.3']
                },
                inplace=True)

        database_cols_projIPG = sorted(database_df.columns[database_slice_start:])
        database_projs = [ x.split('_')[0] for x in database_cols_projIPG if re.search('_', x)]
        database_ips = [ x.split('_')[1] for x in database_cols_projIPG if re.search('_', x)]
        name_db = db.split('/')[-1].split('.')[-2]
        
        
        # uniquify database_projs
        self.database_projIPs = database_cols_projIPG
        self.database_projs = list(dict.fromkeys(database_projs))
        self.database_ips = database_ips

        # set extract columns
        database_cols = [database_df.columns.tolist()[0], database_df.columns.tolist()[2], *database_cols_projIPG]
        
        # extract columns
        database_df_extract = database_df.loc[:, database_cols]

        # Rename BF column labels to map Golden's
        database_df_extract.rename(columns={'Functional Block': 'Block', 'Feature Name': 'Feature'}, inplace=True)

        # 'Block' & 'Feature' DataFrame
        database_df_extract_projs_start = list(database_df_extract.columns).index('Feature') + 1
        self.database_df_extract_BF = database_df_extract.columns[:database_df_extract_projs_start].tolist()

        self.df_extract_projs_start = database_df_extract_projs_start

        # clean up strings in cell values
        for col in self.database_df_extract_BF:
            if self.mode == 'diffgd':
                database_df_extract[col] = (
                    self.feature_names_mapping(
                    self.cleanup(database_df_extract[col]),
                    'db')
                    )
            elif self.mode == 'diffdd':
                database_df_extract[col] = (
                    self.cleanup(database_df_extract[col])
                    )
        
        for col in database_cols_projIPG:
            database_df_extract[col] = (
                self.cleanup(database_df_extract[col])
                )

        # No MultiIndex, sort by values
        database_df_extract_sorted = (database_df_extract.sort_values(
            by=list(database_df_extract.columns)[:database_df_extract_projs_start]))

        # remove index col
        # database_df_extract_sorted_ri = self.index_1st_col(database_df_extract_sorted)

        # display extracted cols for comparison
        self.display_extract_cols(cols=database_cols, source=db)
        logging.info(f"Extracted:\n{database_df_extract.head()}\n...")
        logging.info(f"database projs IP version = {database_cols_projIPG}")
        logging.info(f"databse projs = {database_projs}")

        # dump out the formatted sheet
        self.format(db)

        if self.mode == 'diffgd':
            database_df_extract_sorted.to_excel(self.wr_database_extract_sorted, index=False)
            logging.info(f"Sorted:\n{database_df_extract_sorted.head()}\n...")
            odd_row_coloring(self.wr_database_extract_sorted, "F0F0F0")
            printLog(f"Dumping file {Path(self.cwd, self.wr_database_extract_sorted)}")
            return database_df_extract_sorted
        elif self.mode == 'diffdd':
            self.wr_database_extract_sorted = f"{Path(self.output_dir, f'output_extract_sorted_{name_db}.xlsx')}"
            database_df_extract_sorted.to_excel(self.wr_database_extract_sorted, index=False)
            logging.info(f"Sorted:\n{database_df_extract_sorted.head()}\n...")
            odd_row_coloring(self.wr_database_extract_sorted, "F0F0F0")
            printLog(f"Dumping file {Path(self.cwd, self.wr_database_extract_sorted)}")
            return [database_df_extract_sorted, name_db]


    def process_golden(self, golden: str) -> pd.DataFrame:
        """
        Customized to process `CSME IE OCS Hardware Architecture Features Per Project.xlsm` for comparisons
        """
        golden_df = pd.read_excel(golden, sheet_name='CSE CSME IE OCS Features', skiprows=1, index_col=None)

        # remove `Unnamed: 0` column
        golden_df = golden_df.loc[:, ~golden_df.columns.str.contains('^Unnamed', na=False, regex=True)]
        
        # strip 'Feature' 
        golden_df.loc[1, 2] = golden_df.loc[1, 2].strip()

        # unmerge column 1: Block, CPU, MISA ....
        golden_df[1] = pd.Series(golden_df[1]).fillna(method='ffill')
        
        # unmerge column cells
        golden_df.iloc[0, 3:] = golden_df.iloc[0, 3:].fillna(method='ffill')

        # unmerge Feature cell
        golden_df.loc[0:1, 2] = golden_df.loc[0:1, 2].fillna(method='bfill')

        # Remove footer garbage rows
        golden_df[2] = golden_df[2].fillna(np.nan)
        golden_df = golden_df.dropna(subset=[2])

        # Change golden_df.columns, 1, 2, 3...74, to 'Feature', 'IP Version Number: ', 'OCS 2.0'...
        golden_df.columns = golden_df.iloc[0]

        # set extract columns
        golden_cols = ['Block', 'Feature', *self.database_projs]

        # Extract columns from golden_df
        golden_df_extract = golden_df.loc[:, golden_cols]

        # Remove the duplicate row, column labels
        golden_df_extract = golden_df_extract.drop(index=0)

        # relabel the index
        golden_df_extract.reset_index(drop=True, inplace=True)

        # handle golden extract column labels
        self.golden_extract_projs_start = list(golden_df_extract.columns).index('Feature') + 1
        golden_extract_projs = golden_df_extract.columns[self.golden_extract_projs_start:]
        logging.info('-' * 65)
        logging.info(f"Extracted:\n{golden_df_extract.head()}\n...")

        # Formatting columns, 
        # Replace 'GSC (GSC 3.3)' with 'GSC 3.3' if any
        ipGenerations = golden_df_extract.loc[0, :][self.golden_extract_projs_start:]
        ipGenerations.replace('(.*)\ \((.*)\)', r'\2', regex=True, inplace=True)
        # Replace 'ESE2.3' with 'ESE 2.3' if any
        ipGenerations.replace(r'(\w+)(\d+.\d+)', r'\1 \2', regex=True, inplace=True)

        # Combine proj + ipGeneration as new column labels
        colNames_proj_ipGeneration = ([ '_'.join([proj, ipGeneration]) 
        for proj, ipGeneration in zip(golden_extract_projs, ipGenerations)])

        # projIPs cols to extract
        new_cols_projIPs = []
        for i in colNames_proj_ipGeneration:
            if i in self.database_projIPs:
                new_cols_projIPs.append(i)

        # Replace columns with new labels
        golden_df_extract.columns = (list(golden_df_extract.columns)
        [:self.golden_extract_projs_start] + sorted(colNames_proj_ipGeneration))

        # new extracted cols
        new_cols = (list(golden_df_extract.columns)
        [:self.golden_extract_projs_start] + sorted(new_cols_projIPs)) 
        
        golden_df_extract = golden_df_extract.loc[:, new_cols]

        # display extracted cols for comparison
        self.display_extract_cols(cols=golden_cols, source=golden)
        self.display_extract_cols(cols=self.database_ips, source=golden)
        logging.info(f"Column Names of Proj and IP Generation = {new_cols}" )
        logging.info(f"Rename cols:\n{golden_df_extract.head()}\n...")
        self.display_extract_cols(cols=golden_df_extract.columns, source=golden)
        
        # Remove row 0
        golden_df_extract = golden_df_extract.drop(labels=[0], axis=0)
        logging.info(f"Remove row 0:\n{golden_df_extract.head()}\n...")

        # relabel the index
        golden_df_extract.reset_index(drop=True, inplace=True)

        # clean up column values and then feature names mapping 
        for col in golden_df_extract.columns[:self.golden_extract_projs_start]:
            golden_df_extract[col] = self.feature_names_mapping(self.cleanup(golden_df_extract[col]), 'golden')

        for col in new_cols:
            golden_df_extract[col] = self.cleanup(golden_df_extract[col])

        # remove waived features
        golden_df_extract = self.drop_waived(golden_df_extract)

        # No MultiIndex, sort by values
        golden_df_extract_sorted = (golden_df_extract.sort_values
        (by=list(golden_df_extract.columns)[:self.golden_extract_projs_start]))

        # remove index col
        # golden_df_extract_sorted_ri = self.index_1st_col(golden_df_extract_sorted)

        golden_df_extract_sorted.to_excel(self.wr_golden_extract_sorted, index=False)
        logging.info(f"Sorted:\n{golden_df_extract_sorted.head()}\n...")
        odd_row_coloring(self.wr_golden_extract_sorted, "F0F0F0")
        printLog(f"Dumping file {Path(self.cwd, self.wr_golden_extract_sorted)}")
        return golden_df_extract_sorted
    

    def display_extract_cols(self, cols: list, source: str):
        printLog('Extract ', end='')
        for col in cols:
            printLog(f"'{col}'", end=' ')
        printLog(f'from {source}...')

    def display_compare(self, cols: list):
        printLog('Comparing ', end='')
        for col in cols:
            printLog(f"'{col}'", end=' ')
        printLog('...')

    def drop_waived(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        [09/23/2022]
        These features only exist in Golden but not in ArchGUI database. 
        The team wants to keep these features in the golden sheet.
        Remove them from comparison against database.

        Block	    Feature	    Match
        ------------------------------------- 
        CM DEVICE	IDE-R	    Golden Only
        GASKET	    MROM	    Golden Only
        """
        df_waived = df[df['Block'].isin(['CM DEVICE', 'GASKET'])]
        df_waived = df_waived[df_waived['Feature'].isin(['IDE-R', 'MROM'])]
        df_waived.to_excel(self.wr_waived_golden, index=False)
        printLog(f"Dumping file {Path(self.cwd, self.wr_waived_golden)}")

        df = df.set_index(df.columns[:self.df_extract_projs_start].tolist())
        df.drop(index=('CM DEVICE', 'IDE-R'), inplace=True)
        df.drop(index=('GASKET', 'MROM'), inplace=True)
        df.reset_index(inplace=True)
        return df


    def diff(self, df1, df2):
        """
        Compare two spreadsheets and dumps out files
        """

        mismatch = f"\n[Mismatch Found] - NOT Equivalent; Two objects have mismatched columns\n"
        mismatch += f"For details please check {Path(self.cwd, self.output_dir)}"

        if self.mode == 'diffgd':
            if not df1.columns.tolist() == df2.columns.tolist():
                printLog(mismatch)
                return
            else:
                # diff 1: to dump out discrepancy and the matched
                self.display_compare(cols=df1.columns.tolist())
                discrepancy = (
                    df1.merge(df2, 
                    indicator=True, how='outer')
                    .loc[lambda v: v['_merge'] != 'both']
                    )
                matched = (
                    df1.merge(df2, 
                    indicator=True, 
                    how='outer')
                    .loc[lambda v: v['_merge'] == 'both']
                    )
                # Rename _merge column for discrepancy
                self.rename_merge_col(
                    discrepancy, 
                    'Match', 
                    'golden_only', 
                    'database_only'
                    )
                
                # diff 2: to dump out the extras
                diff_golden = (
                    pd.merge(
                        df1, 
                        df2, 
                        indicator=True, 
                        how='outer')
                        .loc[lambda v: v['_merge'] == 'left_only']
                    )
                diff_database = (
                    pd.merge(
                        df1, 
                        df2, 
                        indicator=True, 
                        how='outer')
                        .loc[lambda v: v['_merge'] == 'right_only']
                    )            
                # Rename _merge column
                self.rename_merge_col(
                    diff_golden, 
                    'Match', 
                    'golden_only', 
                    'database_only'
                    )
                self.rename_merge_col(
                    diff_database, 
                    'Match', 
                    'golden_only', 
                    'database_only'
                    )
                # sort by values
                diff_golden.sort_values(
                    by=diff_golden.columns[self.df_extract_projs_start - 1], 
                    inplace=True
                    )
                diff_database.sort_values(
                    by=diff_database.columns[self.df_extract_projs_start - 1],
                    inplace=True
                    )         
                # remove index col
                # diff_golden_ri = self.index_1st_col(diff_golden)
                # diff_database_ri = self.index_1st_col(diff_database)

                diff_golden.to_excel(self.wr_diff_golden, index=False)
                printLog(f"Dumping file {Path(self.cwd, self.wr_diff_golden)}")
                
                diff_database.to_excel(self.wr_diff_database, index=False)
                printLog(f"Dumping file {Path(self.cwd, self.wr_diff_database)}")

        elif self.mode == 'diffdd':
            if not df1[0].columns.tolist() == df2[0].columns.tolist():
                printLog(mismatch)
                return
            else:
                self.display_compare(cols=df1[0].columns.tolist())
                discrepancy = (
                    df1[0]
                    .merge(df2[0], 
                    indicator=True, how='outer')
                    .loc[lambda v: v['_merge'] != 'both']
                    )
                matched = (
                    df1[0]
                    .merge(df2[0],
                    indicator=True, 
                    how='outer')
                    .loc[lambda v: v['_merge'] == 'both']
                    )            
                # Rename _merge column for discrepancy
                self.rename_merge_col(
                    discrepancy, 
                    'Match', 
                    'database1_only', 
                    'database2_only'
                    )
                
                # diff 2: to dump out the extras
                diff_golden = (
                    pd.merge(
                        df1[0], 
                        df2[0], 
                        indicator=True, 
                        how='outer')
                        .loc[lambda v: v['_merge'] == 'left_only']
                    )
                diff_database = (
                    pd.merge(
                        df1[0], 
                        df2[0], 
                        indicator=True, 
                        how='outer')
                        .loc[lambda v: v['_merge'] == 'right_only']
                    )                 
                # Rename _merge column
                self.rename_merge_col(
                    diff_golden, 
                    'Match', 
                    'database1_only', 
                    'database2_only'
                    )
                self.rename_merge_col(
                    diff_database, 
                    'Match', 
                    'database1_only', 
                    'database2_only'
                    )

                # sort by values
                diff_golden.sort_values(
                    by=diff_golden.columns[self.df_extract_projs_start - 1], 
                    inplace=True
                    )
                diff_database.sort_values(
                    by=diff_database.columns[self.df_extract_projs_start - 1], 
                    inplace=True
                    )                  
                # remove index col
                # diff_golden_ri = self.index_1st_col(diff_golden)
                # diff_database_ri = self.index_1st_col(diff_database)

                self.wr_diff_golden = f"{Path(self.output_dir, f'output_diffcfg_{df1[1]}.xlsx')}"
                diff_golden.to_excel(self.wr_diff_golden, index=False)
                printLog(f"Dumping file {Path(self.cwd, self.wr_diff_golden)}")
                
                self.wr_diff_database = f"{Path(self.output_dir, f'output_diffcfg_{df2[1]}.xlsx')}"
                diff_database.to_excel(self.wr_diff_database, index=False)
                printLog(f"Dumping file {Path(self.cwd, self.wr_diff_database)}")                  

        # Rename _merge column for matched
        self.rename_merge_col(
            matched, 
            'Match', 
            'golden_only', 
            'database_only'
            )
        # sort by values
        discrepancy.sort_values(
            by=discrepancy.columns[self.df_extract_projs_start - 1], 
            inplace=True
        )
        matched.sort_values(
            by=matched.columns.tolist()[:self.df_extract_projs_start],
            inplace=True
        )

        # remove index col
        # discrepancy_ri = self.index_1st_col(discrepancy)
        # matched_ri = self.index_1st_col(matched)

        discrepancy.to_excel(self.wr_discrepancy, index=False)
        printLog(f"Dumping file {Path(self.cwd, self.wr_discrepancy)}")
        matched.to_excel(self.wr_matched, index=False)
        odd_row_coloring(self.wr_matched, "F0F0F0")
        printLog(f"Dumping file {Path(self.cwd, self.wr_matched)}")


        # diff 3: compare 'Block' & 'Feature' columns only
        if self.mode == 'diffgd':
            diffBF_golden = (
                pd.merge(df1.loc[:, self.database_df_extract_BF ], 
                df2.loc[:, self.database_df_extract_BF], 
                indicator=True, 
                how='outer').loc[lambda v: v['_merge'] == 'left_only']
                )
            diffBF_database = (
                pd.merge(df1.loc[:, self.database_df_extract_BF], 
                df2.loc[:, self.database_df_extract_BF], 
                indicator=True, 
                how='outer').loc[lambda v: v['_merge'] == 'right_only']
                )
            diffBF_all = (
                pd.merge(df1.loc[:, self.database_df_extract_BF], 
                df2.loc[:, self.database_df_extract_BF], 
                indicator=True, 
                how='outer').loc[lambda v: v['_merge'] != 'both']
                )
            # Rename _merge column
            self.rename_merge_col(
                diffBF_golden, 
                'Match', 
                'golden_only', 
                'database_only'
                )
            self.rename_merge_col(
                diffBF_database, 
                'Match', 
                'golden_only', 
                'database_only'
                )
            self.rename_merge_col(
                diffBF_all, 
                'Match', 
                'golden_only', 
                'database_only'
                )

            # sort by values
            diffBF_golden.sort_values(
                by=diffBF_golden.columns[self.df_extract_projs_start - 1], 
                inplace=True
            )
            diffBF_database.sort_values(
                by=diffBF_database.columns[self.df_extract_projs_start - 1], 
                inplace=True
            )
            diffBF_all.sort_values(
                by=diffBF_all.columns[self.df_extract_projs_start - 1], 
                inplace=True
            )

            diffBF_golden.to_excel(self.wr_diffBF_golden, index=False)
            printLog(f"Dumping file {Path(self.cwd, self.wr_diffBF_golden)}")

            diffBF_database.to_excel(self.wr_diffBF_database, index=False)
            printLog(f"Dumping file {Path(self.cwd, self.wr_diffBF_database)}")

            diffBF_all.to_excel(self.wr_diffBF_all, index=False)
            printLog(f"Dumping file {Path(self.cwd, self.wr_diffBF_all)}")

            # diffBF_all_ri_html = diffBF_all_ri.to_html()
            # with open('output_diffBF.html', 'w') as f:
            #     f.write(diffBF_all_ri_html)
            # diffBF_all_ri.to_html(f'{self.output_dir}/output_diffBF.html')

            self.equal(df1, df2)

        elif self.mode == 'diffdd':
            diffBF_golden = (
                pd.merge(df1[0].loc[:, self.database_df_extract_BF ], 
                df2[0].loc[:, self.database_df_extract_BF], 
                indicator=True, 
                how='outer').loc[lambda v: v['_merge'] == 'left_only']
                )
            diffBF_database = (
                pd.merge(df1[0].loc[:, self.database_df_extract_BF], 
                df2[0].loc[:, self.database_df_extract_BF], 
                indicator=True, 
                how='outer').loc[lambda v: v['_merge'] == 'right_only']
                )
            diffBF_all = (
                pd.merge(df1[0].loc[:, self.database_df_extract_BF], 
                df2[0].loc[:, self.database_df_extract_BF], 
                indicator=True, 
                how='outer').loc[lambda v: v['_merge'] != 'both']
                )
                
            # Rename _merge column
            self.rename_merge_col(
                diffBF_golden, 
                'Match', 
                'database1_only', 
                'database2_only'
                )
            self.rename_merge_col(
                diffBF_database, 
                'Match', 
                'database1_only', 
                'database2_only'
                )
            self.rename_merge_col(
                diffBF_all, 
                'Match', 
                'database1_only', 
                'database2_only'
                )

            # sort by values
            diffBF_golden.sort_values(
                by=diffBF_golden.columns[self.df_extract_projs_start - 1], 
                inplace=True
            )
            diffBF_database.sort_values(
                by=diffBF_database.columns[self.df_extract_projs_start - 1], 
                inplace=True
            )
            diffBF_all.sort_values(
                by=diffBF_all.columns[self.df_extract_projs_start - 1], 
                inplace=True
            )

            self.wr_diffBF_golden = f"{Path(self.output_dir, f'output_diffbf_extra_{df1[1]}.xlsx')}"
            diffBF_golden.to_excel(self.wr_diffBF_golden, index=False)
            printLog(f"Dumping file {Path(self.cwd, self.wr_diffBF_golden)}")

            self.wr_diffBF_database = f"{Path(self.output_dir, f'output_diffbf_extra_{df2[1]}.xlsx')}"
            diffBF_database.to_excel(self.wr_diffBF_database, index=False)
            printLog(f"Dumping file {Path(self.cwd, self.wr_diffBF_database)}")

            diffBF_all.to_excel(self.wr_diffBF_all, index=False)
            printLog(f"Dumping file {Path(self.cwd, self.wr_diffBF_all)}")

            self.equal(df1[0], df2[0])

    
    def equal(self, df1: pd.DataFrame, df2: pd.DataFrame) -> bool:
        match = f"\n[Matched] - Equivalent\nFor details please check {Path(self.cwd, self.output_dir)}"
        mismatch = f"\n[Mismatch Found] - Mismatched shape: NOT Equivalent\nFor details please check {Path(self.cwd, self.output_dir)}"
        bf_mismatch = f"\n[Mismatch Found] - Mismatches found in Block and Feature columns: NOT Equivalent\nFor details please check {Path(self.cwd, self.output_dir)}"
        ss_mismatch = f"\n[Mismatch Found] - Matched shape: NOT Equivalent\nFor details please check {Path(self.cwd, self.output_dir)}"

        diff = df1.copy()
        equal_BF = df1.loc[:, self.database_df_extract_BF].equals(df2.loc[:, self.database_df_extract_BF])

        # down to this step, df.columns have already been examined to be equivalent already so no need to check it again
        if df1.shape == df2.shape:
            logging.info('df1.shape equals df2.shape...')
            if equal_BF:
                logging.info("df1'BF equals df2's BF...")
                if not df1.equals(df2):
                    df1, df2 = df1.fillna('-'), df2.fillna('-')
                    compareValues = df1.values == df2.values
                    rows, cols = np.where(compareValues == False)
                    for row, col in zip(rows, cols):
                        diff.iloc[row, col] = f"[Diff]: {df1.iloc[row, col]} -> {df2.iloc[row, col]}"
                        
                    diff.to_excel(self.wr_diffSS, index=False)
                    printLog(f"Dumping file {Path(self.cwd, self.wr_diffSS)}")

                    diff_cell_coloring(self.wr_diffSS)
                    printLog(ss_mismatch)
                    return False
                else: 
                    printLog(match)
                    return True
            else:
                logging.info("df1'BF does NOT equal df2's BF...")
                printLog(bf_mismatch)
                return False
        else:
            logging.info('df1.shape does NOT equal df2.shape...')
            printLog(mismatch)
            return False


    def format(self, xlsx: str):
        """
        Formatting xlsx downloaded from databse (archGUI)
        """
        df = pd.read_excel(xlsx)
        df.sort_values(
            by=['Functional Block', 'Feature Name'], 
            inplace=True
        )

        # remove index col
        # df_ri = self.index_1st_col(df)

        # reset index
        # df_rs = df.reset_index(drop=True)
        
        filename = f"output_formatted_{xlsx.split('/')[-1]}"
        dump_file = f"{Path(self.output_dir, filename)}"
        df.to_excel(dump_file, index=False)
        odd_row_coloring(dump_file, "F0F0F0")
        printLog(f"Dumping file {Path(self.cwd, self.output_dir, dump_file)}")

        return df


    def index_1st_col(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        [Deprecated]: use .to_excel(index=False) instead.
        Remove the index col. Set the 1st col index.
        """
        df = df.set_index(df.columns[0], inplace=False)
        return df

    def feature_names_mapping(self, x: pd.Series, source: str) -> pd.Series:
        """
        G: 'PRTC NUM OF PRIVATE CHANNELS\nCHANNEL 0 - ESE\nCHANNEL1 - OSSE'         --> R: 'PRTC NUM OF PRIVATE CHANNELS'
        R: 'DOEMAILBOX'                                                             --> G: 'DOE MAILBOX'
        R: 'FTPM INTERFACE ACCESS TYPE (LT ADDRESS)'                                --> G: 'FTPM INTERFACE ACCESS TYPE'
        R: 'HECI (NUM_INSTANCES)'                                                   --> G: 'HECI'
        G: 'IPC * CHANNEL (DEFAULT *) (16 OR 48 BIT ADDRESSING, DEFAULT IS 16-BIT)' --> R: 'IPC * CHANNEL (DEFAULT *)'
        R: 'PTT (FTPM)'                                                             --> G: 'PTT'        
        R: 'ROOT OF ROOTSPACE'                                                      --> G: 'ROOT OF ROOT SPACE' 
        R: 'TBD (RESERVED GPIC (# WIRES'                                            --> G: 'TBD (RESERVED) GPIC (# WIRES)'         
        R: 'DMA-AES_P (INTEGRITY CHECK VALUE) SCHEME'                               --> G: 'DMA-AES_P ICV (INTEGRITY CHECK VALUE) SCHEME'         
        G: 'OCS  SAVE AND RESTORE'                                                  --> R: 'OCS SAVE AND RESTORE'
        G: '16 BIT IOSF SIDEBAND PORT ID SUPPORT\n(CM DEVICE LIST)'                 --> R: '16 BIT IOSF SIDEBAND PORT ID SUPPORT (CM DEVICE LIST)'
        R: 'GSC FLR(DEVICE RESET)'                                                  --> G: 'GSC FLR (DEVICE RESET)'
        R: 'AES BASIC MODES (ECB,CBC,CTR)'                                          --> G: 'AES BASIC MODES (ECB, CBC, CTR)'
        R: 'AES ADVANCED MODES (OFB,CFB)'                                           --> G: 'AES ADVANCED MODES (OFB, CFB)'
        G: 'L1$ PARITY SUPPORT (TAG )'                                              --> 'L1$ PARITY SUPPORT (TAG)'
        R: 'L1$ PARITY SUPPORT(TAG)'                                                --> 'L1$ PARITY SUPPORT (TAG)'
        R: 'L1$ PARITY SUPPORT(DATA)'                                               --> 'L1$ PARITY SUPPORT (DATA)'
        R: 'BUNIT CACHE SIZE (IN KB)'                                               --> G: 'BUNIT CACHE SIZE'
        R: 'DTF(DEBUG TRACE FABRIC)'                                                --> G: 'DTF (DEBUG TRACE FABRIC)'
        R: 'ECC GEN1 P256(WITH SCA MITIGATION)'                                     --> G: 'ECC GEN1 P256 (WITH SCA MITIGATION)'
        R: 'ECC GEN1 P384(WITH SCA MITIGATION)'                                     --> G: 'ECC GEN1 P384 (WITH SCA MITIGATION)'
        R: 'ECDSA(FW ASSISTED/HW BUILD-IN)'                                         --> G: 'ECDSA (FW ASSISTED / HW BUILD-IN)'
        R: 'HW EXTEND REGISTER FOR FW MEASUREMENT(SHA-256)'                         --> G: 'HW EXTEND REGISTER FOR FW MEASUREMENT (SHA-256)'
        R: 'IOMMU DMA ACCESS CONTROL (# ENTRIES)'                                   --> G: 'IOMMU DMA ACCESS CONTROL'
        R: 'IOMMU TRANSLATION TABLE (# ENTRY)'                                      --> G: 'IOMMU TRANSLATION TABLE'
        R: 'IOSF-P INTERFACE WIDTH( # IN BITS)'                                     --> G: 'IOSF-P INTERFACE WIDTH'
        R: 'L1$ SIZE (CODE + DATA) (# IN KB)'                                       --> G: 'L1$ SIZE (CODE + DATA)'
        R: 'ROM SIZE (# IN KB)'                                                     --> G: 'ROM SIZE'
        R: 'SM4 BASIC MODES(ECB,CBC,CTR)'                                           --> G: 'SM4 BASIC MODES (ECB, CBC, CTR)'
        R: 'SRAM SIZE (EXCLUDING ECC BITS,  # IN KB)'                               --> G: 'SRAM SIZE (EXCLUDING ECC BITS)'
        [09/23/2022]:
        R: '16 BIT IOSF SIDEBAND PORT ID SUPPORT (CSE)'                             --> G: '16 BIT IOSF SIDEBAND PORT ID SUPPORT'
        R: 'GKEY0 (PAVP KEY IN AES-A) (#BITS)'                                      --> G: 'GKEY0 (PAVP KEY IN AES-A)'
        R: 'GKEY3 (# BITS)'                                                         --> G: 'GKEY3's
        G: 'RESET FILTER (RSTFILTER_EN)'                                            --> R: 'RESET FILTER'
        [10/03/2022]:
        G: 'PRTC NUM OF BITS PER CHANNEL (CHANNEL 0)\nWHEN (NUM OF CHANNEL > 0)'    --> R: 'PRTC NUM OF BITS PER CHANNEL (CHANNEL 0) WHEN (NUM OF CHANNEL > 0)'
        G: 'PRTC NUM OF BITS PER CHANNEL (CHANNEL 1)\nWHEN (NUMBER OF CHANNEL > 1)' --> R: 'PRTC NUM OF BITS PER CHANNEL (CHANNEL 1) WHEN (NUM OF CHANNEL > 1)'
        """
        map_list = []
        if x.str.contains('PRTC NUM OF PRIVATE CHANNELS\nCHANNEL 0 - ESE\nCHANNEL 1 - OSSE', regex=False).any():
            map_list.append(['PRTC NUM OF PRIVATE CHANNELS\\nCHANNEL 0 - ESE\\nCHANNEL 1 - OSSE', 'PRTC NUM OF PRIVATE CHANNELS'])
        if x.str.contains('DOEMAILBOX', regex=False).any():
            map_list.append(['DOE MAILBOX', 'DOEMAILBOX'])
        if x.str.contains('FTPM INTERFACE ACCESS TYPE (LT ADDRESS)', regex=False).any():
            map_list.append(['FTPM INTERFACE ACCESS TYPE', 'FTPM INTERFACE ACCESS TYPE (LT ADDRESS)'])
        if x.str.contains('HECI (NUM_INSTANCES)', regex=False).any():
            map_list.append(['HECI', 'HECI (NUM_INSTANCES)'])
        if x.str.contains('IPC * CHANNEL (DEFAULT *) (16 OR 48 BIT ADDRESSING, DEFAULT IS 16-BIT)', regex=False).any():
            map_list.append(['IPC * CHANNEL (DEFAULT *) (16 OR 48 BIT ADDRESSING, DEFAULT IS 16-BIT)', 'IPC * CHANNEL (DEFAULT *)'])
        if x.str.contains('PTT (FTPM)', regex=False).any():
            map_list.append(['PTT', 'PTT (FTPM)'])
        if x.str.contains('ROOT OF ROOTSPACE', regex=False).any():
            map_list.append(['ROOT OF ROOT SPACE', 'ROOT OF ROOTSPACE'])
        if x.str.contains('TBD (RESERVED GPIC (# WIRES', regex=False).any():
            map_list.append(['TBD (RESERVED) GPIC (# WIRES)', 'TBD (RESERVED GPIC (# WIRES'])
        if x.str.contains('DMA-AES_P (INTEGRITY CHECK VALUE) SCHEME', regex=False).any():
            map_list.append(['DMA-AES_P ICV (INTEGRITY CHECK VALUE) SCHEME', 'DMA-AES_P (INTEGRITY CHECK VALUE) SCHEME'])
        if x.str.contains('OCS  SAVE AND RESTORE', regex=False).any():
            map_list.append(['OCS  SAVE AND RESTORE', 'OCS SAVE AND RESTORE'])
        if x.str.contains('16 BIT IOSF SIDEBAND PORT ID SUPPORT\n(CM DEVICE LIST)', regex=False).any():
            map_list.append(['16 BIT IOSF SIDEBAND PORT ID SUPPORT\\n(CM DEVICE LIST)', '16 BIT IOSF SIDEBAND PORT ID SUPPORT (CM DEVICE LIST)'])
        if x.str.contains('GSC FLR(DEVICE RESET)', regex=False).any():
            map_list.append(['GSC FLR (DEVICE RESET)', 'GSC FLR(DEVICE RESET)'])
        if x.str.contains('AES BASIC MODES (ECB,CBC,CTR)', regex=False).any():
            map_list.append(['AES BASIC MODES (ECB, CBC, CTR)', 'AES BASIC MODES (ECB,CBC,CTR)'])
        if x.str.contains('AES ADVANCED MODES (OFB,CFB)', regex=False).any():
            map_list.append(['AES ADVANCED MODES (OFB, CFB)', 'AES ADVANCED MODES (OFB,CFB)'])
        if x.str.contains('L1$ PARITY SUPPORT(TAG)', regex=False).any():
            map_list.append(['L1$ PARITY SUPPORT (TAG)', 'L1$ PARITY SUPPORT(TAG)'])
        if x.str.contains('L1$ PARITY SUPPORT(DATA)', regex=False).any():
            map_list.append(['L1$ PARITY SUPPORT (DATA)', 'L1$ PARITY SUPPORT(DATA)'])
        if x.str.contains('BUNIT CACHE SIZE (IN KB)', regex=False).any():
            map_list.append(['BUNIT CACHE SIZE', 'BUNIT CACHE SIZE (IN KB)'])
        if x.str.contains('DTF(DEBUG TRACE FABRIC)', regex=False).any():
            map_list.append(['DTF (DEBUG TRACE FABRIC)', 'DTF(DEBUG TRACE FABRIC)'])
        if x.str.contains('ECC GEN1 P256(WITH SCA MITIGATION)', regex=False).any():
            map_list.append(['ECC GEN1 P256 (WITH SCA MITIGATION)', 'ECC GEN1 P256(WITH SCA MITIGATION)'])
        if x.str.contains('ECC GEN1 P384(WITH SCA MITIGATION)', regex=False).any():
            map_list.append(['ECC GEN1 P384 (WITH SCA MITIGATION)', 'ECC GEN1 P384(WITH SCA MITIGATION)'])
        if x.str.contains('ECDSA(FW ASSISTED/HW BUILD-IN)', regex=False).any():
            map_list.append(['ECDSA (FW ASSISTED / HW BUILD-IN)', 'ECDSA(FW ASSISTED/HW BUILD-IN)'])
        if x.str.contains('HW EXTEND REGISTER FOR FW MEASUREMENT(SHA-256)', regex=False).any():
            map_list.append(['HW EXTEND REGISTER FOR FW MEASUREMENT (SHA-256)', 'HW EXTEND REGISTER FOR FW MEASUREMENT(SHA-256)'])
        if x.str.contains('IOMMU DMA ACCESS CONTROL (# ENTRIES)', regex=False).any():
            map_list.append(['IOMMU DMA ACCESS CONTROL', 'IOMMU DMA ACCESS CONTROL (# ENTRIES)'])
        if x.str.contains('IOMMU TRANSLATION TABLE (# ENTRY)', regex=False).any():
            map_list.append(['IOMMU TRANSLATION TABLE', 'IOMMU TRANSLATION TABLE (# ENTRY)'])
        if x.str.contains('IOSF-P INTERFACE WIDTH( # IN BITS)', regex=False).any():
            map_list.append(['IOSF-P INTERFACE WIDTH', 'IOSF-P INTERFACE WIDTH( # IN BITS)'])
        if x.str.contains('L1$ SIZE (CODE + DATA) (# IN KB)', regex=False).any():
            map_list.append(['L1$ SIZE (CODE + DATA)', 'L1$ SIZE (CODE + DATA) (# IN KB)'])
        if x.str.contains('ROM SIZE (# IN KB)', regex=False).any():
            map_list.append(['ROM SIZE', 'ROM SIZE (# IN KB)'])
        if x.str.contains('SM4 BASIC MODES(ECB,CBC,CTR)', regex=False).any():
            map_list.append(['SM4 BASIC MODES (ECB, CBC, CTR)', 'SM4 BASIC MODES(ECB,CBC,CTR)'])
        if x.str.contains('SRAM SIZE (EXCLUDING ECC BITS,  # IN KB)', regex=False).any():
            map_list.append(['SRAM SIZE (EXCLUDING ECC BITS)', 'SRAM SIZE (EXCLUDING ECC BITS,  # IN KB)'])
        if x.str.contains('16 BIT IOSF SIDEBAND PORT ID SUPPORT (CSE)', regex=False).any():
            map_list.append(['16 BIT IOSF SIDEBAND PORT ID SUPPORT', '16 BIT IOSF SIDEBAND PORT ID SUPPORT (CSE)'])
        if x.str.contains('GKEY0 (PAVP KEY IN AES-A) (#BITS)', regex=False).any():
            map_list.append(['GKEY0 (PAVP KEY IN AES-A)', 'GKEY0 (PAVP KEY IN AES-A) (#BITS)'])
        if x.str.contains('GKEY3 (# BITS)', regex=False).any():
            map_list.append(['GKEY3', 'GKEY3 (# BITS)'])
        if x.str.contains('RESET FILTER (RSTFILTER_EN)', regex=False).any():
            map_list.append(['RESET FILTER (RSTFILTER_EN)', 'RESET FILTER'])
        if x.str.contains('PRTC NUM OF BITS PER CHANNEL (CHANNEL 0)\nWHEN (NUM OF CHANNEL > 0)', regex=False).any():
            map_list.append(['PRTC NUM OF BITS PER CHANNEL (CHANNEL 0)\\nWHEN (NUM OF CHANNEL > 0)', 'PRTC NUM OF BITS PER CHANNEL (CHANNEL 0) WHEN (NUM OF CHANNEL > 0)'])
        if x.str.contains('PRTC NUM OF BITS PER CHANNEL \(CHANNEL 1\)\nWHEN \(NUMBER OF CHANNEL > 1\)', regex=False).any():
            map_list.append(['PRTC NUM OF BITS PER CHANNEL (CHANNEL 1)\\nWHEN (NUMBER OF CHANNEL > 1)', 'PRTC NUM OF BITS PER CHANNEL (CHANNEL 1) WHEN (NUM OF CHANNEL > 1)'])

        if map_list:
            data_map_list = np.array(map_list)
            df_map_list = pd.DataFrame(data_map_list, columns=['Golden', 'ArchGUI Database'])
            if source == 'golden':
                dump = f"{Path(self.output_dir, f'output_feature_names_mapping_{source}.xlsx')}"
                df_map_list.to_excel(dump, index=False)
                odd_row_coloring(dump, 'F0F0F0')
            elif source == 'db':
                dump = f"{Path(self.output_dir, f'output_feature_names_mapping_{source}.xlsx')}"
                df_map_list.to_excel(dump, index=False)
                odd_row_coloring(dump, 'F0F0F0')

        return (x.replace(to_replace='(PRTC[ ]*NUM[ ]*OF[ ]*PRIVATE[ ]*CHANNELS)[ ]*\n[ ]*(CHANNEL[ ]*0[ ]*-[ ]*ESE)[ ]*\n[ ]*(CHANNEL[ ]*1[ ]*-[ ]*OSSE)',
                value=r'\1', regex=True)                                   
                .replace(to_replace='(DOE)(MAILBOX)',
                value=r'\1 \2', regex=True)
                .replace(to_replace='(FTPM INTERFACE ACCESS TYPE)[ ]*\(LT ADDRESS\)',
                value=r'\1', regex=True)
                .replace(to_replace='HECI[ ]*\(NUM_INSTANCES\)',
                value='HECI', regex=True)
                .replace(to_replace='(IPC \d CHANNEL \(DEFAULT FOR .+\))[ ]*\(16 OR 48 BIT ADDRESSING, DEFAULT IS .+BIT\)',                
                value=r'\1', regex=True)
                .replace(to_replace='(PTT)[ ]*\(FTPM\)',       
                value=r'\1', regex=True)
                .replace(to_replace='(ROOT)[ ]*(OF)[ ]*(ROOT)(SPACE)',
                value=r'\1 \2 \3 \4', regex=True)
                .replace(to_replace='(TBD)[ ]*\((RESERVED)[ ]*(GPIC)[ ]*\([ ]*\#[ ]*(WIRES)',     
                value=r'\1 (\2) \3 (# \4)', regex=True)
                .replace(to_replace='(DMA-AES_P)[ ]*\((INTEGRITY)[ ]*(CHECK)[ ]*(VALUE)[ ]*\)[ ]*(SCHEME)',       
                value=r'\1 ICV (\2 \3 \4) \5', regex=True)
                .replace(to_replace='(OCS)[ ]*(SAVE)[ ]*(AND)[ ]*(RESTORE)',
                value=r'\1 \2 \3 \4', regex=True)
                .replace(to_replace='([ ]*16 BIT IOSF SIDEBAND PORT ID SUPPORT[ ]*)\n\([ ]*(CM DEVICE LIST)[ ]*\)',
                value=r'\1 (\2)', regex=True)
                .replace('(GSC FLR)[ ]*\([ ]*(DEVICE RESET)[ ]*\)', 
                value=r'\1 (\2)', regex=True)
                .replace('(AES BASIC MODES)[ ]*\([ ]*(ECB)[ ]*,[ ]*(CBC)[ ]*,[ ]*(CTR)[ ]*\)',
                value=r'\1 (\2, \3, \4)', regex=True)
                .replace(to_replace='(AES ADVANCED MODES)[ ]*\([ ]*(OFB)[ ]*,[ ]*(CFB)[ ]*\)', 
                value=r'\1 (\2, \3)', regex=True)
                .replace(to_replace='(L1\$ PARITY SUPPORT)[ ]*\([ ]*(\w+)[ ]*\)', 
                value=r'\1 (\2)', regex=True)
                .replace(to_replace='(BUNIT CACHE SIZE)[ ]*\([ ]*IN[ ]*KB[ ]*\)',
                value=r'\1', regex=True)
                .replace(to_replace='(DTF)[ ]*\([ ]*(DEBUG TRACE FABRIC)[ ]*\)',
                value=r'\1 (\2)', regex=True)
                .replace(to_replace='(ECC GEN1 P\d+)[ ]*\([ ]*(WITH SCA MITIGATION)[ ]*\)',
                value=r'\1 (\2)', regex=True)
                .replace(to_replace='(ECDSA)[ ]*\([ ]*(FW)[ ]*(ASSISTED)[ ]*/[ ]*(HW)[ ]*(BUILD-IN)[ ]*\)',                     
                value=r'\1 (\2 \3 / \4 \5)', regex=True)
                .replace(to_replace='(HW EXTEND REGISTER FOR FW MEASUREMENT)[ ]*\([ ]*(SHA-\d+)[ ]*\)',
                value=r'\1 (\2)', regex=True)
                .replace(to_replace='(IOMMU DMA ACCESS CONTROL)[ ]*\([ ]*\#[ ]*ENTRIES[ ]*\)',
                value=r'\1', regex=True)
                .replace(to_replace='(IOMMU TRANSLATION TABLE)[ ]*\([ ]*\#[ ]*ENTRY[ ]*\)',
                value=r'\1', regex=True)
                .replace(to_replace='(IOSF-P INTERFACE WIDTH)[ ]*\([ ]*\#[ ]*IN[ ]*BITS[ ]*\)',
                value=r'\1', regex=True)
                .replace(to_replace='(L1\$ SIZE \(CODE \+ DATA\))[ ]*\([ ]*\#[ ]*IN[ ]*KB[ ]*\)',
                value=r'\1', regex=True)
                .replace(to_replace='(ROM SIZE)[ ]*\([ ]*\#[ ]*IN[ ]*KB[ ]*\)',
                value=r'\1', regex=True)
                .replace(to_replace='(SM4 BASIC MODES)[ ]*\([ ]*(ECB)[ ]*,[ ]*(CBC)[ ]*,[ ]*(CTR)[ ]*\)',
                value=r'\1 (\2, \3, \4)', regex=True)
                .replace(to_replace='(SRAM)[ ]*(SIZE)[ ]*\([ ]*(EXCLUDING)[ ]*(ECC)[ ]*(BITS)[ ]*,[ ]*\#[ ]*IN[ ]*KB[ ]*\)',
                value=r'\1 \2 (\3 \4 \5)', regex=True)
                .replace(to_replace='(16)[ ]*(BIT)[ ]*(IOSF)[ ]*(SIDEBAND)[ ]*(PORT)[ ]*(ID)[ ]*(SUPPORT)[ ]*\([ ]*CSE[ ]*\)',
                value=r'\1 \2 \3 \4 \5 \6 \7', regex=True)
                .replace(to_replace='(GKEY0)[ ]*\([ ]*(PAVP)[ ]*(KEY)[ ]*(IN)[ ]*(AES-A)[ ]*\)[ ]*\([ ]*\#BITS[ ]*\)',
                value=r'\1 (\2 \3 \4 \5)', regex=True)
                .replace(to_replace='(GKEY3)[ ]*\(\#[ ]*BITS\)',
                value=r'\1', regex=True)
                .replace(to_replace='(RESET)[ ]*(FILTER)[ ]*\([ ]*RSTFILTER_EN[ ]*\)',
                value=r'\1 \2', regex=True)
                .replace(to_replace='(PRTC[ ]*NUM*[ ]*OF[ ]*BITS[ ]*PER[ ]*CHANNEL[ ]*\([ ]*CHANNEL[ ]*\d[ ]*\))[ ]*\n[ ]*(WHEN)[ ]*\(NUM[BER]*[ ]*(OF[ ]*CHANNEL[ ]*>[ ]*\d\))',
                value=r'\1 \2 (NUM \3', regex=True)
                )

    def cleanup(self, x):
        """
        Capitalize and strip to clean up the formatting of strings that cause false discrepancy
        """
        return (x.replace(to_replace='(Yes)[ ]*[\n]*[ ]*\([ ]*(x3,100,12.8)[ ]*\)', value=r'\1 (\2)', regex=True)  # For 'Yes \n(x3,100,12.8)'
                .apply(lambda x: x.strip().upper() if isinstance(x, str) else x))
    
    def rename_merge_col(self, df: pd.DataFrame, label, left, right):
        """
        rename _merge col label and values 
        """
        df.rename(
            columns={df.columns[len(df.columns) - 1]: label}, 
            inplace=True)
        df[df.columns[len(df.columns) - 1]] = (
            df[df.columns[len(df.columns) - 1]]
            .replace(to_replace='left_only', value=left)
            .replace(to_replace='right_only', value=right))


def func_diffxlsx(df1: pd.DataFrame, df2: pd.DataFrame, outdir: str) -> bool:
    match = f"\n[Matched] - Equivalent"
    mismatch = f"\n[Mismatch Found] - Mismatched shape: NOT Equivalent"
    cols_mismatch = f"\n[Mismatch Found] - Mismatched columns: NOT Equivalent"
    if df1.columns.tolist() == df2.columns.tolist():
        logging.info('df1 columns equal df2 columns...')
        if df1.shape == df2.shape:
            logging.info('df1 shape equals df2 shape...')
            df1_sorted, df2_sorted = df1.sort_values(by=df1.columns.tolist()), df2.sort_values(by=df2.columns.tolist())
            df1_sorted, df2_sorted = df1_sorted.reset_index(drop=True), df2_sorted.reset_index(drop=True)

            if not df1_sorted.equals(df2_sorted):
                p = Path(outdir)
                p.mkdir(parents=True, exist_ok=True)
                dump_file = f"{Path(outdir, 'output_diffxlsx.xlsx')}"
                cwd = Path.cwd()

                ss_mismatch = f"\n[Mismatch Found] - Matched shape: NOT Equivalent\nFor details please check {Path(cwd, outdir)}"
                diff = df1_sorted.copy()

                df1_sorted, df2_sorted = df1_sorted.fillna('-'), df2_sorted.fillna('-')
                compareValues = df1_sorted.values == df2_sorted.values
                rows, cols = np.where(compareValues == False)

                for row, col in zip(rows, cols):
                    diff.iloc[row, col] = f"[Diff]: {df1_sorted.iloc[row, col]} -> {df2_sorted.iloc[row, col]}"

                diff.reset_index(drop=True, inplace=True)
                # To remove `Unnamed: 0` column
                diff = diff.drop(columns=diff.columns[0])
                diff.to_excel(dump_file, index=False)
                printLog(f"Dumping file {Path(cwd, dump_file)}...")
                odd_row_coloring(dump_file, 'F0F0F0')
                diff_cell_coloring(dump_file)
                printLog(ss_mismatch)
                return False
            else: 
                printLog(match)
                return True
        else:
            logging.info('df1 shape does NOT equal df2 shape...')
            printLog(mismatch)
            return False
    else:
        logging.info('df1 columns does NOT equal df2 columns...')
        printLog(cols_mismatch)
        return False


def diff_cell_coloring(xlsx: str):
    """
    [Diff] cell formatting, filled with bgcolor yellow
    """
    wb = load_workbook(xlsx)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if re.search("^\[Diff\].*$", str(cell_value)):
                ws.cell(row, col).fill = PatternFill(fill_type='solid',
                                            start_color='FFFF00',
                                            end_color='FFFF00')
    wb.save(xlsx)


def odd_row_coloring(xlsx: str, color: str):
    """
    odd-row formatting, filled with bgcolor grey
    """
    wb = load_workbook(xlsx)
    ws = wb['Sheet1']
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if row == 1:
                ws.cell(row, col).fill = PatternFill(fill_type="solid",
                                                start_color='FFFFFF',
                                                end_color='FFFFFF')
            elif row % 2:
                ws.cell(row, col).fill = PatternFill(start_color=color, 
                                                end_color=color, 
                                                fill_type="solid")
    wb.save(xlsx)


def setup_parser(script_name):
    """
    Set up the argument parser
    """
    descript = "Customized for processing 'CSME IE OCS Hardware Architecture Features Per Project.xlsm'"

    parser = aps.ArgumentParser(
        prog=f'{script_name}.py',
        description=descript
    )
    subparsers = parser.add_subparsers(dest='command')
    subparsers.required = True

    # subparser for diffgd
    parser_diffgd = subparsers.add_parser(
        'diffgd',
        help="To compare two HW arch features config sheets: 'CSME IE OCS Hardware Architecture Features Per Project.xlsm' VS. HW arch features config from ArchGUI database")
    # add required args
    parser_diffgd.add_argument(
        '-g',
        '--golden',
        type=str,
        metavar='',
        default='CSME IE OCS Hardware Architecture Features Per Project.xlsm',
        help="'CSME IE OCS Hardware Architecture Features Per Project.xlsm'"
    )
    parser_diffgd.add_argument(
        '-d',
        '--database',
        type=str,
        metavar='',
        required=True,
        help='HW arch features config file from ArchGUI database'         
    )
    parser_diffgd.add_argument(
        '-o',
        '--out_dir',
        type=str,
        metavar='',
        default=f'{script_name}_outputs',
        help=f'Output directory, default "{script_name}_outputs"'
    )

    # subparser for diffdd
    parser_diffdd = subparsers.add_parser(
        'diffdd',
        help="To compare two HW arch features config sheets, both from ArchGUI database")
    # add required args
    parser_diffdd.add_argument(
        '-f1',
        '--file1',
        type=str,
        metavar='',
        required=True,
        help='HW arch features config file1 from ArchGUI database'
    )
    parser_diffdd.add_argument(
        '-f2',
        '--file2',
        type=str,
        metavar='',
        required=True,
        help='HW arch features config file2 from ArchGUI database'
    )
    parser_diffdd.add_argument(
        '-o',
        '--out_dir',
        type=str,
        metavar='',
        default=f'{script_name}_outputs',
        help=f'Output directory, default "{script_name}_outputs"'
    )

    # subparser for diffxlsx
    parser_diffxlsx = subparsers.add_parser(
        'diffxlsx',
        help='A quick checker to test whether two xlsx files have the same shape and the same elements, Only dump out a comparison report when two objects are in the same shape but NOT Equivalent')
    # add required args
    parser_diffxlsx.add_argument(
        '-f',
        '--files',
        type=str,
        metavar='',
        nargs=2,
        required=True,
        help='Two xlsx files for comparison'
    )
    parser_diffxlsx.add_argument(
        '-o',
        '--out_dir',
        type=str,
        metavar='',
        default=f'{script_name}_outputs',
        help=f'Output directory, default "{script_name}_outputs". Only dump out a comparison report when two objects are in the same shape but have discrepancy'
    )

    # subparser for format
    parser_format = subparsers.add_parser(
        'format',
        help='To format HW arch features config sheet from ArchGUI database')
    # add a required arg
    parser_format.add_argument(
        '-f',
        '--file',
        type=str,
        metavar='',
        required=True,
        help='HW arch features config file from ArchGUI database'
    )
    parser_format.add_argument(
        '-o',
        '--out_dir',
        type=str,
        metavar='',
        default=f'{script_name}_outputs',
        help=f'Output directory, default "{script_name}_outputs"'
    )
    return parser

def printLog(msg, end='\n'):
    print(msg, end=end)
    logging.info(msg)

def main():
    script = 'archqa'
    run_log = 'archqa.run.log'
    parser = setup_parser(script)
    args = parser.parse_args()
    curr_time = datetime.now().strftime('[%m/%d/%Y %a %I:%M%p]')
    dashes = '-' * 9
    done = f'\n{curr_time}\n{dashes} DONE {dashes}'

    p = Path(args.out_dir)
    p.mkdir(parents=True, exist_ok=True)
    
    logging.basicConfig(
        filename=f"{Path(args.out_dir,run_log)}", 
        level=logging.DEBUG,
        format='%(message)s',
        # format='%(asctime)s %(message)s',
        datefmt='[%m/%d/%Y %a %I:%M%p]',
        filemode='w'
    )
    # logger = logging.getLogger()
    # logger.setLevel(logging.DEBUG)

    def display():
        print()
        for k, v in vars(args).items():
            printLog(f"{str.upper(k):>8} : {v}")
        print()
    
    if args.command == 'diffgd':
        display()
        d = XlsxDiff(args.out_dir, args.command)
        df_db = d.process_database(args.database)
        df_golden = d.process_golden(args.golden)
        d.diff(df_golden, df_db)
        printLog(done)
    elif args.command == 'diffdd':
        display()
        d = XlsxDiff(args.out_dir, args.command)
        df_db1 = d.process_database(args.file1)
        df_db2 = d.process_database(args.file2)
        d.diff(df_db1, df_db2)
        printLog(done)
    elif args.command == 'format':
        display()
        d = XlsxDiff(args.out_dir, args.command)
        d.format(args.file)
        printLog(done)
    elif args.command == 'diffxlsx':
        display()
        file1, file2, out = pd.read_excel(args.files[0]), pd.read_excel(args.files[1]), args.out_dir
        func_diffxlsx(file1, file2, out)
        printLog(done)

if __name__ == "__main__":
    main()
        
